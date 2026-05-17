import os
import random
import time
import string
import smtplib
import csv
import io
from datetime import datetime, timedelta
from email.mime.text import MIMEText

try:
    import mysql.connector as mysql
except ImportError:
    import pymysql as mysql

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    psycopg2 = None
    RealDictCursor = None

from flask import Flask, redirect, render_template, request, session, flash, jsonify
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from python.dashboard_routes import build_dashboard_context
from python.interview_service import InterviewService

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# =========================
# INIT
# =========================
# Get absolute path to .env file
basedir = os.path.abspath(os.path.dirname(__file__))
load_dotenv(os.path.join(basedir, '.env'))

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY")

# Configure Upload Folder
if os.environ.get('VERCEL'):
    UPLOAD_FOLDER = '/tmp/uploads'
else:
    UPLOAD_FOLDER = os.path.join(app.root_path, 'static', 'uploads')

try:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'profiles'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'documents'), exist_ok=True)
except OSError as e:
    print(f"Warning: Could not create upload directories: {e}")

def handle_file_upload(file, roletype, name):
    if not file or file.filename == '':
        return None
    
    # Naming convention: roletype_name.ext
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
    # Sanitize name
    clean_name = "".join([c if c.isalnum() else "_" for c in name.lower()]).strip("_")
    filename = secure_filename(f"{roletype}_{clean_name}_{int(time.time())}.{ext}")
    
    upload_path = os.path.join(UPLOAD_FOLDER, 'profiles', filename)
    file.save(upload_path)
    return f"/static/uploads/profiles/{filename}"

# =========================
# DATABASE CONNECTION
# =========================
def is_postgres_enabled():
    return bool(os.getenv("DATABASE_URL") or os.getenv("SUPABASE_DATABASE_URL") or os.getenv("SUPABASE_DB_URL") or os.getenv("POSTGRES_URL"))


def get_db():
    if is_postgres_enabled():
        if psycopg2 is None:
            raise RuntimeError("psycopg2 is not installed. PostgreSQL database operations are unavailable.")

        database_url = (
            os.getenv("DATABASE_URL")
            or os.getenv("SUPABASE_DATABASE_URL")
            or os.getenv("SUPABASE_DB_URL")
            or os.getenv("POSTGRES_URL")
        )
        connect_kwargs = {"sslmode": os.getenv("POSTGRES_SSLMODE", "require")}
        if "sslmode=" in database_url:
            connect_kwargs.pop("sslmode", None)
        return psycopg2.connect(database_url, **connect_kwargs)

    if not mysql:
        raise RuntimeError("MySQL connector not installed. Database operations are unavailable.")
    return mysql.connect(
        host=os.getenv("DB_HOST") or os.getenv("MYSQL_HOST", "localhost"),
        user=os.getenv("DB_USER") or os.getenv("MYSQL_USER", "root"),
        password=os.getenv("DB_PASSWORD") if os.getenv("DB_PASSWORD") is not None else os.getenv("MYSQL_PASSWORD", ""),
        database=os.getenv("DB_NAME") or os.getenv("MYSQL_DATABASE"),
        port=int(os.getenv("DB_PORT") or os.getenv("MYSQL_PORT", "3306")),
    )

def get_db_cursor(db, dictionary=False):
    if is_postgres_enabled():
        if dictionary:
            return db.cursor(cursor_factory=RealDictCursor)
        return db.cursor()

    if dictionary:
        try:
            return db.cursor(dictionary=True)
        except TypeError:
            # Fallback for PyMySQL
            try:
                import pymysql.cursors
                return db.cursor(pymysql.cursors.DictCursor)
            except ImportError:
                return db.cursor()
    return db.cursor()


def execute_insert_returning_id(db, cursor, sql, params):
    if is_postgres_enabled():
        sql = sql.rstrip().rstrip(";")
        if "returning" not in sql.lower():
            sql = f"{sql} RETURNING id"
        cursor.execute(sql, params)
        inserted = cursor.fetchone()
        if isinstance(inserted, dict):
            return inserted.get("id")
        if inserted:
            return inserted[0]
        return None

    cursor.execute(sql, params)
    return cursor.lastrowid

def get_user_settings(user_id):
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("SELECT * FROM user_settings WHERE user_id = %s", (user_id,))
    settings = cursor.fetchone()
    
    if not settings:
        # Create default settings if not exists
        cursor.execute("INSERT INTO user_settings (user_id) VALUES (%s)", (user_id,))
        db.commit()
        cursor.execute("SELECT * FROM user_settings WHERE user_id = %s", (user_id,))
        settings = cursor.fetchone()
        
    db.close()
    return settings

def get_global_setting(key, default=None):
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT setting_value FROM settings WHERE setting_key = %s", (key,))
        result = cursor.fetchone()
        db.close()
        return result[0] if result else default
    except Exception:
        return default

def get_notification_data(user_id, role=None):
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        if role == 'admin':
            # Admin sees global notifications
            cursor.execute("SELECT COUNT(*) as total FROM notifications WHERE is_read = 0")
            count = cursor.fetchone()["total"]
            cursor.execute("SELECT * FROM notifications ORDER BY created_at DESC LIMIT 5")
            recent = cursor.fetchall()
        else:
            # Users see their own notifications
            cursor.execute("SELECT COUNT(*) as total FROM notifications WHERE user_id = %s AND is_read = 0", (user_id,))
            count = cursor.fetchone()["total"]
            cursor.execute("SELECT * FROM notifications WHERE user_id = %s ORDER BY created_at DESC LIMIT 5", (user_id,))
            recent = cursor.fetchall()
        db.close()
        
        ui_notes = []
        for n in recent:
            ui_notes.append({
                "icon": "solar:bell-linear" if n["type"] == "system" else "solar:shield-star-bold-duotone",
                "title": n["title"],
                "subtitle": n["created_at"].strftime("%b %d"),
                "tone": "text-primary" if n["type"] == "system" else "text-blue-500"
            })
            
        return count, ui_notes
    except Exception:
        return 0, []

# =========================
# ACTIVITY LOGGING
# =========================
def log_activity(user_id, action, details=None, activity_type='admin'):
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            INSERT INTO activities (user_id, action, details, type, ip_address, user_agent)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (user_id, action, details, activity_type, request.remote_addr if request else None, request.user_agent.string if request else None))
        db.commit()
        db.close()
    except Exception as e:
        print(f"Failed to log activity: {e}")

def refresh_session_data():
    if "user_id" in session:
        count, ui_notes = get_notification_data(session["user_id"], session.get("role"))
        session["notification_count"] = count
        session["ui_notifications"] = ui_notes

# =========================
# EMAIL FUNCTION (BREVO)
# =========================
def send_email(to_email, code):
    try:
        host = os.getenv("BREVO_SMTP_HOST")
        port_str = os.getenv("BREVO_SMTP_PORT")
        user = os.getenv("BREVO_SMTP_USER")
        password = os.getenv("BREVO_SMTP_PASSWORD")
        sender = os.getenv("BREVO_SENDER_EMAIL")

        if not all([host, port_str, user, password, sender]):
            print("Email configuration is missing in .env")
            return False

        port = int(port_str)
        
        msg = MIMEText(f"Your password reset code is: {code}\nValid for 10 minutes.")
        msg["Subject"] = "Reset Password - AI Interviewer"
        msg["From"] = sender
        msg["To"] = to_email

        server = smtplib.SMTP(host, port, timeout=10)
        server.starttls()
        server.login(user, password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        # Use a safe way to log error that won't cause OSError 22
        try:
            print(f"Email error: {str(e)}")
        except OSError:
            pass
        return False

# =========================
# ROLE REDIRECT
# =========================
def role_redirect(role):
    if role == "admin":
        return redirect("/admin/dashboard")
    elif role == "company":
        return redirect("/companies/dashboard")
    elif role == "client":
        return redirect("/client/dashboard")
    elif role == "company_user":
        return redirect("/client/interview")
    return redirect("/login")

# =========================
# LOGIN
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("auth/login.html")

    email = request.form.get("email")
    password = request.form.get("password")

    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)

    cursor.execute("SELECT * FROM users WHERE email=%s", (email,))
    user = cursor.fetchone()

    if not user:
        return render_template("auth/login.html", error="Invalid email or password")

    # Check if user account is active
    if not user["is_active"]:
        return render_template("auth/login.html", error="Your account has been suspended. Please contact support for assistance.")

    # For company_users, check if their company admin account is active
    if user["role"] == "company_user" and user["company_id"]:
        cursor.execute("""
            SELECT u.is_active FROM users u 
            JOIN companies c ON c.email = u.email 
            WHERE c.id = %s AND u.role = 'company'
        """, (user["company_id"],))
        company_admin = cursor.fetchone()
        if company_admin and not company_admin["is_active"]:
            return render_template("auth/login.html", error="Your company account has been suspended. Please contact your company administrator.")

    try:
        if not check_password_hash(user["password_hash"], password):
            return render_template("auth/login.html", error="Invalid email or password")
    except ValueError:
        # Invalid hash method, treat as invalid password
        return render_template("auth/login.html", error="Invalid email or password")

    session["user_id"] = user["id"]
    session["role"] = user["role"]
    session["email"] = user["email"]
    session["full_name"] = f"{user['first_name']} {user['last_name']}"
    session["profile_image"] = user.get("profile_img")

    refresh_session_data()

    return role_redirect(user["role"])

# =========================
# SIGNUP
# =========================
@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "GET":
        return render_template("auth/signup.html")

    full_name = request.form.get("full_name")
    email = request.form.get("email")
    password = request.form.get("password")
    confirm = request.form.get("confirm_password")
    role = "client"

    if password != confirm:
        return render_template("auth/signup.html", error="Passwords do not match")

    first_name = full_name.split(" ")[0]
    last_name = " ".join(full_name.split()[1:]) if len(full_name.split()) > 1 else ""

    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)

    cursor.execute("SELECT id FROM users WHERE email=%s", (email,))
    if cursor.fetchone():
        db.close()
        return render_template("auth/signup.html", error="Email already exists")

    hashed = generate_password_hash(password, method='scrypt')

    new_user_id = execute_insert_returning_id(db, cursor, """
        INSERT INTO users (email, password_hash, first_name, last_name, role)
        VALUES (%s, %s, %s, %s, %s)
    """, (email, hashed, first_name, last_name, role))
    
    # Auto-assign free subscription plan for client
    cursor.execute("""
        SELECT id, monthly_price FROM subscription_plans 
        WHERE target_audience='individual' AND monthly_price=0 
        LIMIT 1
    """)
    free_plan = cursor.fetchone()
    
    if free_plan:
        cursor.execute("""
            INSERT INTO subscriptions 
            (user_id, plan_id, status, billing_cycle, price, start_date, end_date)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (new_user_id, free_plan['id'], 'active', 'monthly', free_plan['monthly_price'],
              datetime.now().date(), 
              (datetime.now() + timedelta(days=30)).date()))

    db.commit()
    db.close()

    flash("Signup successful! Please login.")
    return redirect("/login")

# =====================================================
# 🔐 FORGOT PASSWORD (3 STEP IN ONE PAGE)
# =====================================================
@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():

    step = session.get("step", 1)

    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)

    if request.method == "POST":

        # ---------- STEP 1: EMAIL ----------
        if step == 1:
            email = request.form.get("email")

            cursor.execute("SELECT id FROM users WHERE email=%s", (email,))
            user = cursor.fetchone()

            if not user:
                return render_template("auth/forgot-password.html", error="Email not found", step=1)

            code = ''.join(random.choices(string.digits, k=6))
            expires = datetime.now() + timedelta(minutes=10)

            cursor.execute("DELETE FROM password_reset_codes WHERE user_id=%s", (user["id"],))

            cursor.execute("""
                INSERT INTO password_reset_codes (user_id, code, expires_at)
                VALUES (%s, %s, %s)
            """, (user["id"], code, expires))

            db.commit()

            if send_email(email, code):
                session["reset_email"] = email
                session["step"] = 2
                return render_template("auth/forgot-password.html", step=2)
            else:
                return render_template("auth/forgot-password.html", error="Failed to send reset code. Please try again later.", step=1)

        # ---------- STEP 2: VERIFY CODE ----------
        elif step == 2:
            code = request.form.get("code")
            email = session.get("reset_email")

            cursor.execute("SELECT id FROM users WHERE email=%s", (email,))
            user = cursor.fetchone()

            cursor.execute("""
                SELECT * FROM password_reset_codes
                WHERE user_id=%s AND code=%s
                ORDER BY created_at DESC LIMIT 1
            """, (user["id"], code))

            record = cursor.fetchone()

            if not record:
                return render_template("auth/forgot-password.html", error="Invalid code", step=2)

            if record["expires_at"] < datetime.now():
                return render_template("auth/forgot-password.html", error="Code expired", step=2)

            session["reset_user_id"] = user["id"]
            session["step"] = 3

            return render_template("auth/forgot-password.html", step=3)

        # ---------- STEP 3: RESET PASSWORD ----------
        elif step == 3:
            password = request.form.get("password")
            confirm = request.form.get("confirm_password")

            if password != confirm:
                return render_template("auth/forgot-password.html", error="Passwords do not match", step=3)

            user_id = session.get("reset_user_id")

            cursor.execute("""
                UPDATE users SET password_hash=%s WHERE id=%s
            """, (generate_password_hash(password), user_id))

            cursor.execute("DELETE FROM password_reset_codes WHERE user_id=%s", (user_id,))
            db.commit()

            session.clear()

            flash("Password reset successful!")
            return redirect("/login")

    return render_template("auth/forgot-password.html", step=step)

# =========================
# Admin
# =========================
@app.route("/admin/dashboard")
def admin_dashboard():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
    
    refresh_session_data()
    
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    
    # Platform stats
    cursor.execute("SELECT COUNT(*) AS total FROM users")
    total_users = cursor.fetchone()["total"]
    
    cursor.execute("SELECT COUNT(*) AS total FROM companies")
    total_companies = cursor.fetchone()["total"]
    
    cursor.execute("SELECT SUM(price) AS total FROM subscriptions WHERE status='active'")
    total_revenue = cursor.fetchone()["total"] or 0
    
    cursor.execute("SELECT COUNT(*) AS total FROM interviews WHERE status='completed'")
    total_interviews = cursor.fetchone()["total"]
    
    # Recent data for tables
    cursor.execute("SELECT * FROM users ORDER BY created_at DESC LIMIT 5")
    recent_users = cursor.fetchall()
    
    cursor.execute("""
        SELECT c.*, 
        (SELECT COUNT(*) FROM users u WHERE u.company_id = c.id) as team_size
        FROM companies c 
        ORDER BY created_at DESC LIMIT 5
    """)
    recent_companies = cursor.fetchall()
    
    cursor.execute("""
        SELECT t.*, u.email as user_email, u.role as user_role
        FROM support_tickets t
        JOIN users u ON u.id = t.user_id
        ORDER BY t.created_at DESC LIMIT 5
    """)
    recent_tickets = cursor.fetchall()
    
    cursor.execute("""
        SELECT a.*, u.email as user_email
        FROM activities a
        LEFT JOIN users u ON u.id = a.user_id
        ORDER BY a.created_at DESC LIMIT 10
    """)
    recent_activities = cursor.fetchall()
    
    db.close()
    
    context = build_dashboard_context("admin", "dashboard")
    stats = {
        "total_users": total_users,
        "total_companies": total_companies,
        "total_revenue": total_revenue,
        "total_interviews": total_interviews
    }
    
    context["stats"] = stats
    context["recent_users"] = recent_users
    context["recent_companies"] = recent_companies
    context["recent_tickets"] = recent_tickets
    context["recent_activities"] = recent_activities
    
    return render_template("admin/dashboard.html", **context)

@app.route("/admin/activities")
def admin_activities():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
    context = build_dashboard_context("admin", "activities")
    return render_template("admin/activities.html", **context)

@app.route("/api/logs")
@app.route("/read_logs")
def read_logs():
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    log_type = request.args.get("type", "all").strip().lower()
    limit_raw = request.args.get("limit", "150")

    try:
        limit = max(1, min(int(limit_raw), 300))
    except (TypeError, ValueError):
        limit = 150

    if log_type not in {"all", "admin", "company", "client", "system"}:
        log_type = "all"

    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)

    where_clause = ""
    params = []
    if log_type != "all":
        where_clause = "WHERE a.type = %s"
        params.append(log_type)

    cursor.execute(f"""
        SELECT
            a.id,
            a.action,
            a.details,
            a.type AS log_type,
            a.ip_address,
            a.user_agent,
            a.created_at,
            u.email AS user_email,
            NULL AS path,
            NULL AS status_code
        FROM activities a
        LEFT JOIN users u ON u.id = a.user_id
        {where_clause}
        ORDER BY a.created_at DESC, a.id DESC
        LIMIT %s
    """, (*params, limit))
    logs = cursor.fetchall()

    cursor.execute("""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN type = 'admin' THEN 1 ELSE 0 END) AS admin,
            SUM(CASE WHEN type = 'company' THEN 1 ELSE 0 END) AS company,
            SUM(CASE WHEN type = 'client' THEN 1 ELSE 0 END) AS client,
            SUM(CASE WHEN type = 'system' THEN 1 ELSE 0 END) AS system
        FROM activities
    """)
    counts = cursor.fetchone() or {}

    db.close()

    return jsonify({
        "logs": logs,
        "counts": counts,
        "count": len(logs),
        "filter": log_type,
    }), 200

@app.route("/admin/company")
def admin_company():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
    context = build_dashboard_context("admin", "company")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("""
        SELECT
            c.id AS platform_id,
            c.id,
            c.company_name,
            c.first_name,
            c.last_name,
            c.email,
            c.phone,
            c.country,
            c.region,
            c.address,
            c.profile_img,
            c.created_at,
            u.cnic AS cnic,
            COALESCE(u.is_active, 1) AS is_active,
            COALESCE((SELECT COUNT(*) FROM users cu WHERE cu.company_id = c.id AND cu.role = 'company_user'), 0) AS team_size,
            COALESCE((
                SELECT s.plan_id
                FROM users company_user
                JOIN subscriptions s ON s.user_id = company_user.id AND s.status = 'active'
                WHERE company_user.email = c.email AND company_user.role = 'company'
                ORDER BY s.created_at DESC
                LIMIT 1
            ), '') AS current_plan_id
        FROM companies c
        LEFT JOIN users u ON u.email = c.email AND u.role = 'company'
        ORDER BY c.created_at DESC, c.id DESC
    """)
    companies = cursor.fetchall()

    cursor.execute("""
        SELECT
            d.id AS doc_id,
            d.company_id,
            d.file_name AS document_name,
            d.file_path,
            'pending' AS status,
            d.uploaded_at,
            c.company_name
        FROM company_documents d
        JOIN companies c ON c.id = d.company_id
        ORDER BY d.uploaded_at DESC, d.id DESC
    """)
    documents = cursor.fetchall()
    db.close()

    documents_by_company = {}
    for document in documents:
        company_id = int(document.get("company_id") or 0)
        documents_by_company.setdefault(company_id, []).append(document)

    companies_with_documents = []
    for company in companies:
        company["id"] = company.get("platform_id")
        company["user_count"] = company.get("team_size", 0)
        company["documents"] = documents_by_company.get(int(company.get("platform_id") or 0), [])
        companies_with_documents.append(company)

    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("SELECT id, name FROM subscription_plans WHERE target_audience = 'company'")
    plans = cursor.fetchall()
    
    cursor.execute("SELECT * FROM contact_sales ORDER BY created_at DESC")
    contact_sales = cursor.fetchall()
    db.close()

    return render_template(
        "admin/company.html",
        **context,
        companies=companies,
        admin_companies_table=companies,
        admin_company_documents=documents,
        admin_companies_with_documents=companies_with_documents,
        subscription_plans=plans,
        contact_sales=contact_sales
    )

@app.route("/admin/interview-management")
def admin_interview_management():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
    
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("""
        SELECT i.*, u.first_name, u.last_name, u.email as user_email,
        c.company_name
        FROM interviews i
        JOIN users u ON u.id = i.user_id
        LEFT JOIN companies c ON c.id = i.company_id
        ORDER BY i.created_at DESC
    """)
    interviews = cursor.fetchall()
    
    cursor.execute("SELECT id, first_name, last_name, profile_img FROM users WHERE role = 'client'")
    users = cursor.fetchall()
    
    db.close()
    
    context = build_dashboard_context("admin", "interview-management")
    context["interviews"] = interviews
    context["users"] = users
    return render_template("admin/interview-management.html", **context)

@app.route("/admin/notification/delete/<int:id>", methods=["DELETE"])
def delete_admin_notification(id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("DELETE FROM notifications WHERE id = %s", (id,))
        db.commit()
        db.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/admin/notification")
def admin_notification():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("SELECT * FROM notifications ORDER BY created_at DESC")
    notifications = cursor.fetchall()
    
    cursor.execute("SELECT id, username, first_name, last_name, role FROM users WHERE role IN ('client', 'company_user')")
    all_users = cursor.fetchall()
    
    cursor.execute("SELECT id, company_name, email FROM companies")
    all_companies = cursor.fetchall()
    
    db.close()
    
    context = build_dashboard_context("admin", "notification")
    return render_template("admin/notification.html", 
        **context, 
        notifications=notifications,
        all_users=all_users,
        all_companies=all_companies
    )

@app.route("/admin/profile")
def admin_profile():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("SELECT * FROM users WHERE id = %s", (session["user_id"],))
    user = cursor.fetchone()
    db.close()
    
    context = build_dashboard_context("admin", "profile")
    return render_template("admin/profile.html", **context, user=user)

@app.route("/admin/settings")
def admin_settings():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
        
    context = build_dashboard_context("admin", "settings")
    user_settings = get_user_settings(session.get("user_id"))
    
    # Fetch global settings
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("SELECT * FROM settings ORDER BY setting_group")
    global_settings = cursor.fetchall()
    db.close()
    
    return render_template("admin/settings.html", **context, user_settings=user_settings, global_settings=global_settings)

@app.route("/admin/subscription")
def admin_subscription():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("SELECT * FROM subscription_plans ORDER BY id")
    plans = cursor.fetchall()
    
    cursor.execute("SELECT id, company_name FROM companies")
    companies = cursor.fetchall()
    db.close()
    
    context = build_dashboard_context("admin", "subscription")
    return render_template("admin/subscription.html", **context, plans=plans, companies=companies)

@app.route("/admin/support")
def admin_support():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("""
        SELECT t.*, u.email as user_email, u.role as user_role, u.first_name, u.last_name
        FROM support_tickets t
        JOIN users u ON u.id = t.user_id
        ORDER BY t.created_at DESC
    """)
    tickets = cursor.fetchall()
    db.close()
    
    context = build_dashboard_context("admin", "support")
    return render_template("admin/support.html", **context, tickets=tickets)

@app.route("/admin/users")
def admin_users():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/login")
    context = build_dashboard_context("admin", "users")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("""
        SELECT
            u.id,
            u.username,
            u.email,
            u.role,
            u.is_active,
            u.created_at,
            u.first_name,
            u.last_name,
            u.phone,
            u.profile_img
        FROM users u
        ORDER BY u.created_at DESC, u.id DESC
    """)
    users = cursor.fetchall()
    db.close()
    return render_template("admin/users.html", **context, users=users, admin_users=users)

# =========================
# Company
# =========================

@app.route("/companies/dashboard")
def company_dashboard():
    # Ensure user is logged in and belongs to a company
    if "user_id" not in session or session.get("role") != "company":
        return redirect("/login")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)

    # Get company id for the logged-in company
    cursor.execute("SELECT id FROM companies WHERE email = %s LIMIT 1", (session.get("email"),))
    comp = cursor.fetchone()
    company_id = comp["id"] if comp else None
    if not company_id:
        db.close()
        return redirect("/login")

    # 1. Total company users
    cursor.execute("SELECT COUNT(*) AS total FROM users WHERE company_id = %s AND role = 'company_user'", (company_id,))
    total_candidates = cursor.fetchone()["total"]

    # 2. Active interviews for this company
    cursor.execute("SELECT COUNT(*) AS active FROM interviews WHERE company_id = %s AND status = 'in_progress'", (company_id,))
    active_interviews = cursor.fetchone()["active"]

    # 3. Average clearance rating for this company's completed interviews
    cursor.execute("SELECT AVG(overall_score) AS avg_score FROM interviews WHERE company_id = %s AND status = 'completed'", (company_id,))
    avg_score = cursor.fetchone()["avg_score"] or 0

    # 4. Security checks rate – placeholder 100%
    security_rate = 100

    # 5. Recent company users – include employee_id from company_users mapping
    cursor.execute("""
        SELECT u.id, u.email, u.first_name, u.last_name, u.designation, cu.employee_id
        FROM users u
        LEFT JOIN company_users cu ON cu.user_id = u.id
        WHERE u.company_id = %s AND u.role = 'company_user'
        ORDER BY u.created_at DESC LIMIT 5
    """, (company_id,))
    recent_company_users = cursor.fetchall()

    db.close()

    stats = {
        "total_candidates": total_candidates,
        "active_interviews": active_interviews,
        "avg_score": round(avg_score, 2),
        "security_rate": security_rate,
    }

    context = build_dashboard_context("companies", "dashboard")
    context["stats"] = stats
    context["recent_company_users"] = recent_company_users
    return render_template("companies/dashboard.html", **context)

@app.route("/companies/interview-management")
def company_interview_management():
    if "user_id" not in session or session.get("role") != "company":
        return redirect("/login")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    
    # Get company ID
    cursor.execute("SELECT id FROM companies WHERE email = %s LIMIT 1", (session.get("email"),))
    comp = cursor.fetchone()
    company_id = comp["id"] if comp else None

    # Fetch company users
    company_users = []
    if company_id:
            cursor.execute("""
                SELECT u.id, u.first_name, u.last_name, u.designation, u.email, cu.employee_id
                FROM users u
                LEFT JOIN company_users cu ON cu.user_id = u.id
                WHERE u.company_id=%s AND u.role='company_user' AND u.designation IS NOT NULL AND u.designation != ''
                ORDER BY u.first_name, u.last_name
            """, (company_id,))
            company_users = cursor.fetchall()

    # Fetch upcoming interviews assigned to this company (include pending + in_progress)
    cursor.execute("""
        SELECT i.*, u.first_name, u.last_name, u.email AS user_email, cu.employee_id AS user_employee_id
        FROM interviews i
        LEFT JOIN users u ON u.id = i.user_id
        LEFT JOIN company_users cu ON cu.user_id = u.id AND cu.company_id = i.company_id
        WHERE i.company_id = %s AND i.status IN ('pending', 'in_progress')
        ORDER BY i.created_at DESC
    """, (company_id,))
    upcoming = cursor.fetchall()
    if not upcoming:
        # small debug hint for developers: no interviews found for this company
        print(f"[debug] No upcoming interviews for company_id={company_id}")
    else:
        # compute completed counts grouped by role for this company to show evaluated numbers
        cursor.execute("SELECT target_role, COUNT(*) AS completed_count FROM interviews WHERE company_id = %s AND status = 'completed' GROUP BY target_role", (company_id,))
        completed_rows = cursor.fetchall()
        completed_map = {r['target_role']: r['completed_count'] for r in completed_rows}
        for iv in upcoming:
            iv['evaluated_count'] = completed_map.get(iv.get('target_role'), 0)
    # pass count for debugging and log fetched rows
    upcoming_count = len(upcoming)
    print(f"[debug] upcoming_count={upcoming_count} for company_id={company_id}; upcoming_ids={[i.get('id') for i in upcoming]}")
    # Fetch active API providers (models)
    cursor.execute("SELECT DISTINCT provider FROM api_keys WHERE status = 'active'")
    active_providers = [p['provider'] for p in cursor.fetchall()]

    db.close()
    
    context = build_dashboard_context("companies", "interview-management")
    return render_template("companies/interview-management.html", upcoming_interviews=upcoming, company_users=company_users, active_models=active_providers, company_id=company_id, upcoming_count=upcoming_count, **context)

@app.route("/companies/interview/delete/<int:id>", methods=["DELETE"])
def delete_company_interview(id):
    if "user_id" not in session or session.get("role") != "company":
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Verify the interview belongs to this company
        cursor.execute("SELECT id FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
        comp = cursor.fetchone()
        if not comp:
            return jsonify({"success": False, "error": "Company not found"}), 404
            
        cursor.execute("DELETE FROM interviews WHERE id = %s AND company_id = %s", (id, comp[0]))
        db.commit()
        db.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/companies/interview/deploy", methods=["POST"])
def deploy_interview():
    if "user_id" not in session or session.get("role") != "company":
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    try:
        data = request.json
        designation = data.get("designation")
        selected_users = data.get("selected_users", [])
        duration = data.get("duration", 30)
        language = data.get("language", "English")
        
        if not designation or not selected_users:
            return jsonify({"success": False, "error": "Missing required fields"}), 400
            
        db = get_db()
        cursor = db.cursor()
        
        # Get company ID
        cursor.execute("SELECT id FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
        comp = cursor.fetchone()
        if not comp:
            return jsonify({"success": False, "error": "Company not found"}), 404
            
        company_id = comp[0]
        
        for identifier in selected_users:
            # allow either numeric user_id or 8-char employee_id
            user_id = None
            try:
                user_id = int(identifier)
            except Exception:
                # treat as employee_id -> lookup user_id
                cursor.execute("SELECT user_id FROM company_users WHERE employee_id = %s AND company_id = %s LIMIT 1", (identifier, company_id))
                row = cursor.fetchone()
                if row:
                    user_id = row[0]
                else:
                    # invalid identifier, skip
                    continue

            # Lookup employee_id for the resolved user_id
            cursor.execute("SELECT employee_id FROM company_users WHERE user_id = %s AND company_id = %s LIMIT 1", (user_id, company_id))
            emp = cursor.fetchone()
            employee_id = emp[0] if emp else None

            cursor.execute("""
                INSERT INTO interviews (user_id, target_role, interview_type, duration, status, company_id, voice_language, employee_id)
                VALUES (%s, %s, 'company', %s, 'pending', %s, %s, %s)
            """, (user_id, designation, duration, company_id, language, employee_id))
            
        db.commit()
        db.close()
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/companies/interview/results")
def company_interview_results():
    if "user_id" not in session or session.get("role") != "company":
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    target_role = request.args.get("role")
    if not target_role:
        return jsonify({"success": False, "error": "Missing role parameter"}), 400
        
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        cursor.execute("SELECT id FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
        comp = cursor.fetchone()
        if not comp:
            return jsonify({"success": False, "error": "Company not found"}), 404
            
        cursor.execute("""
            SELECT i.id, i.overall_score, i.status, u.first_name, u.last_name, u.email
            FROM interviews i
            JOIN users u ON u.id = i.user_id
            WHERE i.company_id = %s AND i.target_role = %s
            ORDER BY i.overall_score DESC, i.created_at DESC
        """, (comp["id"], target_role))
        
        results = cursor.fetchall()
        db.close()
        
        return jsonify({"success": True, "results": results})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/companies/notification")
def company_notification():
    if "user_id" not in session or session.get("role") != "company":
        return redirect("/login")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    
    # Get company ID
    cursor.execute("SELECT id FROM companies WHERE email = %s LIMIT 1", (session.get("email"),))
    comp = cursor.fetchone()
    company_id = comp["id"] if comp else None
    
    if not company_id:
        db.close()
        return redirect("/login")

    cursor.execute("""
        SELECT n.* FROM notifications n
        JOIN users u ON u.id = n.user_id
        WHERE u.company_id = %s
        ORDER BY n.created_at DESC
    """, (company_id,))
    notes = cursor.fetchall()
    
    # Fetch all users of this company for the Compose modal
    cursor.execute("SELECT id, first_name, last_name, designation FROM users WHERE company_id = %s", (company_id,))
    company_users = cursor.fetchall()
    
    db.close()
    context = build_dashboard_context("companies", "notification")
    return render_template("companies/notification.html", notifications=notes, company_users=company_users, **context)

@app.route("/companies/profile")
def company_profile():
    # Fetch company details for dynamic profile page
    if "user_id" not in session:
        return redirect("/login")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    company = None
    # If logged in as company role, fetch by email
    if session.get("role") == "company":
        cursor.execute("SELECT * FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
        company = cursor.fetchone()
    else:
        # Fetch company_id from user record
        cursor.execute("SELECT company_id FROM users WHERE id=%s LIMIT 1", (session["user_id"],))
        user = cursor.fetchone()
        if user and user.get("company_id"):
            cursor.execute("SELECT * FROM companies WHERE id=%s LIMIT 1", (user["company_id"],))
            company = cursor.fetchone()
    db.close()
    if not company:
        # Fallback mock data if not found or DB unavailable
        company = {
            "first_name": "Company",
            "last_name": "Admin",
            "profile_img": None,
            "account_type": "Company Account",
            "email": "admin@company.com",
            "company_name": "MedxAnalysis Tech",
        }
    context = build_dashboard_context("companies", "profile")
    return render_template("companies/profile.html", company=company, **context)

@app.route("/companies/settings")
def company_settings():
    if "user_id" not in session or session.get("role") != "company":
        return redirect("/login")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    # Get company details
    cursor.execute("SELECT * FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
    company = cursor.fetchone()
    db.close()
    if not company:
        return redirect("/companies/dashboard")
    context = build_dashboard_context("companies", "settings")
    user_settings = get_user_settings(session.get("user_id"))
    return render_template("companies/settings.html", company=company, user_settings=user_settings, **context)

@app.route("/companies/subscription")
def company_subscription():
    if "user_id" not in session:
        return redirect("/login")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    # For company users, perhaps fetch based on company_id, but for now, similar to client
    cursor.execute("""
        SELECT s.*, p.name, p.monthly_price, p.yearly_price, p.features
        FROM subscriptions s
        JOIN subscription_plans p ON s.plan_id = p.id
        WHERE s.user_id = %s AND s.status = 'active'
        ORDER BY s.id DESC LIMIT 1
    """, (session["user_id"],))
    subscription = cursor.fetchone()
    
    cursor.execute("SELECT * FROM subscription_plans WHERE target_audience = 'company' AND is_active = 1 ORDER BY monthly_price ASC")
    plans = cursor.fetchall()
    
    cursor.execute("SELECT * FROM payment_methods WHERE user_id = %s ORDER BY is_primary DESC, created_at DESC", (session["user_id"],))
    payment_methods = cursor.fetchall()
    db.close()
    
    # Parse features for plans
    import json
    for plan in plans:
        if plan.get('features') and isinstance(plan['features'], str):
            try:
                plan['parsed_features'] = json.loads(plan['features'])
            except:
                plan['parsed_features'] = [f.strip() for f in plan['features'].split(',')]
        else:
            plan['parsed_features'] = plan.get('features', [])
            
    # Mask card numbers for template
    for pm in payment_methods:
        cnum = pm['card_number'].replace(' ', '')
        if len(cnum) >= 7:
            pm['masked_card'] = f"{cnum[:3]} •••• •••• {cnum[-4:]}"
        else:
            pm['masked_card'] = "•••• •••• ••••"
            
    context = build_dashboard_context("companies", "subscription")
    return render_template("companies/subscription.html", subscription=subscription, plans=plans, payment_methods=payment_methods, **context)

@app.route("/companies/support")
def company_support():
    if "user_id" not in session or session.get("role") != "company":
        return redirect("/login")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    # Fetch support tickets for this company (via users linked to company)
    cursor.execute("""
        SELECT t.* FROM support_tickets t
        JOIN users u ON u.id = t.user_id
        WHERE u.company_id = (SELECT id FROM companies WHERE email=%s LIMIT 1)
        ORDER BY t.created_at DESC
    """, (session.get("email"),))
    tickets = cursor.fetchall()
    # Fetch current user details for the support forms
    cursor.execute("SELECT first_name, last_name, email FROM users WHERE id = %s LIMIT 1", (session["user_id"],))
    user_data = cursor.fetchone() or {"first_name": "User", "last_name": "", "email": session.get("email", "")}
    db.close()
    
    context = build_dashboard_context("companies", "support")
    return render_template("companies/support.html", tickets=tickets, user=user_data, **context)

@app.route("/api/payment_methods", methods=["POST"])
def add_payment_method():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 403
    
    name_on_card = request.form.get("name_on_card")
    card_number = request.form.get("card_number")
    expiry = request.form.get("expiry")
    cvv = request.form.get("cvv")
    
    if not all([name_on_card, card_number, expiry, cvv]):
        return jsonify({"error": "Missing required fields"}), 400
        
    db = get_db()
    cursor = get_db_cursor(db)
    
    # Check if this is the first card, make it primary if so
    cursor.execute("SELECT COUNT(*) FROM payment_methods WHERE user_id=%s", (session["user_id"],))
    count = cursor.fetchone()[0]
    is_primary = 1 if count == 0 else 0
    
    cursor.execute("""
        INSERT INTO payment_methods (user_id, name_on_card, card_number, expiry, cvv, is_primary)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (session["user_id"], name_on_card, card_number, expiry, cvv, is_primary))
    
    db.commit()
    db.close()
    return jsonify({"success": "Card added successfully"}), 201

@app.route("/api/payment_methods/set_primary", methods=["POST"])
def set_primary_payment_method():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 403
        
    card_id = request.form.get("card_id")
    if not card_id:
        return jsonify({"error": "Missing card_id"}), 400
        
    db = get_db()
    cursor = get_db_cursor(db)
    
    cursor.execute("UPDATE payment_methods SET is_primary=0 WHERE user_id=%s", (session["user_id"],))
    cursor.execute("UPDATE payment_methods SET is_primary=1 WHERE id=%s AND user_id=%s", (card_id, session["user_id"]))
    
    db.commit()
    db.close()
    return jsonify({"success": "Primary card updated"}), 200

@app.route("/api/subscription/upgrade", methods=["POST"])
def upgrade_subscription():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 403
        
    plan_name = request.form.get("plan_name")
    billing_cycle = request.form.get("billing_cycle", "monthly")
    
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    
    # Check for primary payment method
    cursor.execute("SELECT id FROM payment_methods WHERE user_id=%s AND is_primary=1", (session["user_id"],))
    pm = cursor.fetchone()
    if not pm:
        db.close()
        return jsonify({"error": "No primary payment method. Please add a card first."}), 400
        
    # Get plan details
    cursor.execute("SELECT id, monthly_price, yearly_price FROM subscription_plans WHERE name=%s", (plan_name,))
    plan = cursor.fetchone()
    if not plan:
        db.close()
        return jsonify({"error": "Plan not found"}), 404
        
    price = plan["monthly_price"] if billing_cycle == 'monthly' else plan["yearly_price"]
    
    # Deactivate current subscription
    cursor.execute("UPDATE subscriptions SET status='inactive' WHERE user_id=%s AND status='active'", (session["user_id"],))
    
    # Create new subscription
    cursor.execute("""
        INSERT INTO subscriptions (user_id, plan_id, status, billing_cycle, price, start_date, end_date)
        VALUES (%s, %s, 'active', %s, %s, %s, %s)
    """, (session["user_id"], plan["id"], billing_cycle, price, datetime.now().date(), (datetime.now() + timedelta(days=30 if billing_cycle=='monthly' else 365)).date()))
    
    db.commit()
    db.close()
    
    return jsonify({"success": f"Successfully updated subscription to {plan_name}"}), 200

@app.route("/companies/users")
def company_users():
    if "user_id" not in session or session.get("role") != "company":
        return redirect("/login")
    context = build_dashboard_context("companies", "users")
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    # Get company ID of logged-in company user
    cursor.execute("SELECT id FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
    comp = cursor.fetchone()
    company_id = comp["id"] if comp else None
    if not company_id:
        db.close()
        return redirect("/login")
    cursor.execute("""
        SELECT
            u.id,
            u.first_name,
            u.last_name,
            u.designation,
            u.phone,
            u.cnic,
            u.profile_img,
            u.username,
            u.email,
            u.is_active,
            u.created_at
        FROM users u
        WHERE u.company_id=%s AND u.role='company_user'
        ORDER BY u.created_at DESC, u.id DESC
    """, (company_id,))
    users = cursor.fetchall()
    db.close()
    return render_template("companies/users.html", **context, users=users, team_members=users)

# =========================
# clients
# =========================

@app.route("/client/dashboard")
def client_dashboard():
    if "user_id" not in session or session.get("role") not in ["client", "company_user"]:
        return redirect("/login")
        
    refresh_session_data()

    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)

    cursor.execute("SELECT COUNT(*) as total FROM interviews WHERE user_id = %s", (session["user_id"],))
    total_interviews = cursor.fetchone()["total"]
    
    cursor.execute("SELECT AVG(overall_score) as avg FROM interviews WHERE user_id = %s AND status = 'completed'", (session["user_id"],))
    avg_score = cursor.fetchone()["avg"] or 0
    
    cursor.execute("SELECT COUNT(*) as pending FROM interviews WHERE user_id = %s AND status = 'in_progress'", (session["user_id"],))
    pending_interviews = cursor.fetchone()["pending"]
    
    # 2. Recent Interviews
    cursor.execute("SELECT * FROM interviews WHERE user_id = %s ORDER BY created_at DESC LIMIT 5", (session["user_id"],))
    recent_interviews = cursor.fetchall()
    
    db.close()
    
    stats = {
        "total_interviews": total_interviews,
        "avg_score": round(avg_score, 1),
        "pending": pending_interviews
    }
    
    context = build_dashboard_context("client", "dashboard")
    context["stats"] = stats
    context["recent_interviews"] = recent_interviews
    return render_template("client/dashboard.html", **context)

@app.route("/client/interview")
def client_interview():
    if "user_id" not in session or session.get("role") not in ["client", "company_user"]:
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)

    interviews = []
    company_interviews = []
    company_role = ""

    if session.get("role") == "company_user":
        cursor.execute("SELECT company_id FROM users WHERE id = %s", (session["user_id"],))
        user = cursor.fetchone()
        if user and user.get("company_id"):
            cursor.execute(
                "SELECT * FROM interviews WHERE company_id = %s AND interview_type = 'company' ORDER BY created_at DESC",
                (user["company_id"],),
            )
            company_interviews = cursor.fetchall()
            if company_interviews:
                company_role = company_interviews[0].get("target_role") or "Company Interview"
    else:
        cursor.execute("SELECT * FROM interviews WHERE user_id = %s ORDER BY created_at DESC", (session["user_id"],))
        interviews = cursor.fetchall()

    db.close()
    
    context = build_dashboard_context("client" if session.get("role") == "client" else session.get("role", "client"), "interview")
    return render_template(
        "client/interview.html",
        **context,
        interviews=interviews,
        company_interviews=company_interviews,
        company_role=company_role,
    )

@app.route("/client/notification")
def client_notification():
    if "user_id" not in session or session.get("role") not in ["client", "company_user"]:
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    
    cursor.execute("""
        SELECT * FROM notifications 
        WHERE user_id = %s OR type IN ('admin', 'system')
        ORDER BY created_at DESC
    """, (session["user_id"],))
    notifications = cursor.fetchall()
    db.close()
    
    context = build_dashboard_context("client" if session.get("role") == "client" else session.get("role", "client"), "notification")
    return render_template("client/notification.html", **context, db_notifications=notifications)

@app.route("/client/profile")
def client_profile():
    if "user_id" not in session or session.get("role") not in ["client", "company_user"]:
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("""
        SELECT u.*, cu.employee_id 
        FROM users u 
        LEFT JOIN company_users cu ON u.id = cu.user_id 
        WHERE u.id = %s
    """, (session["user_id"],))
    user = cursor.fetchone()
    db.close()
    
    context = build_dashboard_context("client" if session.get("role") == "client" else session.get("role", "client"), "profile")
    return render_template("client/profile.html", **context, user=user)

@app.route("/client/reports")
def client_reports():
    if "user_id" not in session or session.get("role") not in ["client"]:
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("SELECT * FROM interviews WHERE user_id = %s AND status = 'completed' ORDER BY created_at DESC", (session["user_id"],))
    reports = cursor.fetchall()
    db.close()
    
    context = build_dashboard_context("client", "reports")
    return render_template("client/reports.html", **context, reports=reports)

@app.route("/client/settings")
def client_settings():
    if "user_id" not in session or session.get("role") not in ["client"]:
        return redirect("/login")
        
    context = build_dashboard_context("client", "settings")
    user_settings = get_user_settings(session.get("user_id"))
    return render_template("client/settings.html", **context, user_settings=user_settings)

@app.route("/client/subscription")
def client_subscription():
    if "user_id" not in session or session.get("role") not in ["client"]:
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("""
        SELECT s.*, p.name as plan_name, p.monthly_price, p.yearly_price, p.features 
        FROM subscriptions s 
        JOIN subscription_plans p ON s.plan_id = p.id 
        WHERE s.user_id = %s AND s.status = 'active'
    """, (session["user_id"],))
    subscription = cursor.fetchone()
    db.close()
    
    context = build_dashboard_context("client", "subscription")
    return render_template("client/subscription.html", **context, subscription=subscription)

@app.route("/client/support")
def client_support():
    if "user_id" not in session or session.get("role") not in ["client", "company_user"]:
        return redirect("/login")
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("SELECT * FROM support_tickets WHERE user_id = %s ORDER BY created_at DESC", (session["user_id"],))
    tickets = cursor.fetchall()

    # Fetch current user details for the support forms
    cursor.execute("SELECT first_name, last_name, email FROM users WHERE id = %s LIMIT 1", (session["user_id"],))
    user_data = cursor.fetchone() or {"first_name": "User", "last_name": "", "email": session.get("email", "")}
    db.close()
    
    role = session.get("role", "client")
    context = build_dashboard_context(role, "support")
    return render_template("client/support.html", **context, tickets=tickets, user=user_data)

# =========================
# MAIN PAGES
# =========================
@app.route("/")
def index():
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    cursor.execute("SELECT * FROM subscription_plans WHERE is_active = 1 ORDER BY monthly_price ASC")
    plans = cursor.fetchall()
    db.close()
    return render_template("index.html", subscription_plans=plans)

@app.route("/about-us")
def about_us():
    return render_template("pages/about-us.html")

@app.route("/contact-us")
def contact_us():
    return render_template("pages/contact-us.html")

@app.route("/faq")
def faq():
    return render_template("pages/faq.html")

@app.route("/privacy-policy")
def privacy_policy():
    return render_template("pages/privacy-policy.html")

@app.route("/term-and-condition")
def term_and_condition():
    return render_template("pages/term-and-condition.html")

@app.route("/contact-sales", methods=["GET", "POST"])
def contact_sales():
    if request.method == "POST":
        name = request.form.get("name")
        email = request.form.get("email")
        subject = request.form.get("subject")
        message = request.form.get("message")
        
        db = get_db()
        cursor = get_db_cursor(db)
        cursor.execute(
            "INSERT INTO contact_sales (name, email, subject, message) VALUES (%s, %s, %s, %s)",
            (name, email, subject, message)
        )
        db.commit()
        db.close()
        
        flash("Your request has been submitted successfully.", "success")
        return redirect("/contact-sales")
        
    return render_template("pages/contact-sales.html")

# =========================
# USER CREATION & MANAGEMENT
# =========================
@app.route("/admin/users/create", methods=["POST"])
def admin_create_user():
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        first_name = request.form.get("first_name")
        last_name = request.form.get("last_name")
        email = request.form.get("email")
        password = request.form.get("password")
        user_type = request.form.get("user_type")  # 'client' or 'company'
        subscription_plan_id = request.form.get("subscription_plan_id")  # For company type
        
        if not all([first_name, last_name, email, password, user_type]):
            return jsonify({"error": "All fields required"}), 400
        
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if user exists
        cursor.execute("SELECT id FROM users WHERE email=%s", (email,))
        if cursor.fetchone():
            db.close()
            return jsonify({"error": "Email already exists"}), 400
        
        hashed_password = generate_password_hash(password, method='scrypt')
        role = user_type  # 'client' or 'company'
        
        new_user_id = execute_insert_returning_id(db, cursor, """
            INSERT INTO users (first_name, last_name, email, password_hash, role)
            VALUES (%s, %s, %s, %s, %s)
        """, (first_name, last_name, email, hashed_password, role))
        
        # Auto-assign subscription plan
        if user_type == 'client':
            # Get free plan for individual
            cursor.execute("""
                SELECT id FROM subscription_plans 
                WHERE target_audience='individual' AND monthly_price=0 
                LIMIT 1
            """)
            plan = cursor.fetchone()
            plan_id = plan['id'] if plan else None
        elif user_type == 'company':
            # Use selected subscription plan
            plan_id = subscription_plan_id
        else:
            plan_id = None
        
        # Create subscription if plan_id exists
        if plan_id:
            cursor.execute("""
                SELECT monthly_price, yearly_price FROM subscription_plans WHERE id=%s
            """, (plan_id,))
            plan = cursor.fetchone()
            
            if plan:
                cursor.execute("""
                    INSERT INTO subscriptions 
                    (user_id, plan_id, status, billing_cycle, price, start_date, end_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (new_user_id, plan_id, 'active', 'monthly', plan['monthly_price'], 
                      datetime.now().date(), 
                      (datetime.now() + timedelta(days=30)).date()))
        
        db.commit()
        db.close()
        
        return jsonify({"success": "User created successfully"}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/users/update/<int:user_id>", methods=["POST"])
def admin_update_user(user_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        first_name = request.form.get("first_name")
        last_name = request.form.get("last_name")
        email = request.form.get("email")
        phone = request.form.get("phone")
        role = request.form.get("role")
        is_active = request.form.get("is_active") == "true"
        
        if not all([first_name, last_name, email, role]):
            return jsonify({"error": "Required fields missing"}), 400
        
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if user exists
        cursor.execute("SELECT id FROM users WHERE id=%s", (user_id,))
        if not cursor.fetchone():
            db.close()
            return jsonify({"error": "User not found"}), 404

        # Check if email is already used by another user
        cursor.execute("SELECT id FROM users WHERE email=%s AND id!=%s", (email, user_id))
        if cursor.fetchone():
            db.close()
            return jsonify({"error": "Email already used by another user"}), 400

        # Update user
        cursor.execute("""
            UPDATE users 
            SET first_name=%s, last_name=%s, email=%s, phone=%s, role=%s, is_active=%s
            WHERE id=%s
        """, (first_name, last_name, email, phone, role, is_active, user_id))
        
        db.commit()
        db.close()
        
        log_activity(session["user_id"], f"Updated user {first_name} {last_name}", f"User ID: {user_id}")
        return jsonify({"success": "User updated successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/users/suspend/<int:user_id>", methods=["POST"])
def admin_suspend_user(user_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if user exists
        cursor.execute("SELECT id, first_name, last_name FROM users WHERE id=%s", (user_id,))
        user = cursor.fetchone()
        if not user:
            db.close()
            return jsonify({"error": "User not found"}), 404

        # Suspend user
        cursor.execute("UPDATE users SET is_active=0 WHERE id=%s", (user_id,))
        
        db.commit()
        db.close()
        
        log_activity(session["user_id"], f"Suspended user {user['first_name']} {user['last_name']}", f"User ID: {user_id}")
        return jsonify({"success": "User suspended successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/users/unsuspend/<int:user_id>", methods=["POST"])
def admin_unsuspend_user(user_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if user exists
        cursor.execute("SELECT id, first_name, last_name FROM users WHERE id=%s", (user_id,))
        user = cursor.fetchone()
        if not user:
            db.close()
            return jsonify({"error": "User not found"}), 404

        # Unsuspend user
        cursor.execute("UPDATE users SET is_active=1 WHERE id=%s", (user_id,))
        
        db.commit()
        db.close()
        
        log_activity(session["user_id"], f"Unsuspended user {user['first_name']} {user['last_name']}", f"User ID: {user_id}")
        return jsonify({"success": "User unsuspended successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/users/delete/<int:user_id>", methods=["DELETE"])
def admin_delete_user(user_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if user exists
        cursor.execute("SELECT id, first_name, last_name FROM users WHERE id=%s", (user_id,))
        user = cursor.fetchone()
        if not user:
            db.close()
            return jsonify({"error": "User not found"}), 404

        # Delete user
        cursor.execute("DELETE FROM users WHERE id=%s", (user_id,))
        
        db.commit()
        db.close()
        
        log_activity(session["user_id"], f"Deleted user {user['first_name']} {user['last_name']}", f"User ID: {user_id}")
        return jsonify({"success": "User deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/users/import", methods=["POST"])
def admin_import_users():
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        file = request.files.get("file")
        user_type = request.form.get("user_type", "client")  # 'client' or 'company'
        
        if not file:
            return jsonify({"error": "No file provided"}), 400
        
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        created_count = 0
        errors = []
        
        # Get free plan for individuals
        cursor.execute("""
            SELECT id, monthly_price FROM subscription_plans 
            WHERE target_audience='individual' AND monthly_price=0 
            LIMIT 1
        """)
        free_plan = cursor.fetchone()
        free_plan_id = free_plan['id'] if free_plan else None
        
        if file_ext == 'csv':
            stream = io.TextIOWrapper(file.stream, encoding='utf-8')
            reader = csv.DictReader(stream)
            cursor_insert = db.cursor()
            
            for row_num, row in enumerate(reader, start=2):
                try:
                    first_name = row.get('first_name', '').strip()
                    last_name = row.get('last_name', '').strip()
                    email = row.get('email', '').strip()
                    password = row.get('password', '').strip() or os.urandom(12).hex()
                    
                    if not all([first_name, last_name, email]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        continue
                    
                    cursor_insert.execute("SELECT id FROM users WHERE email=%s", (email,))
                    if cursor_insert.fetchone():
                        errors.append(f"Row {row_num}: Email already exists")
                        continue
                    
                    hashed_password = generate_password_hash(password, method='scrypt')
                    
                    new_user_id = execute_insert_returning_id(db, cursor_insert, """
                        INSERT INTO users (first_name, last_name, email, password_hash, role)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (first_name, last_name, email, hashed_password, user_type))
                    
                    # Auto-assign free plan for clients
                    if user_type == 'client' and free_plan_id:
                        cursor_insert.execute("""
                            INSERT INTO subscriptions 
                            (user_id, plan_id, status, billing_cycle, price, start_date, end_date)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (new_user_id, free_plan_id, 'active', 'monthly', 0, 
                              datetime.now().date(), 
                              (datetime.now() + timedelta(days=30)).date()))
                    
                    created_count += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
        
        elif file_ext == 'xlsx' and load_workbook:
            workbook = load_workbook(file.stream)
            worksheet = workbook.active
            cursor_insert = db.cursor()
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    first_name = str(row[0] or '').strip()
                    last_name = str(row[1] or '').strip()
                    email = str(row[2] or '').strip()
                    password = str(row[3] or '').strip() or os.urandom(12).hex()
                    
                    if not all([first_name, last_name, email]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        continue
                    
                    cursor_insert.execute("SELECT id FROM users WHERE email=%s", (email,))
                    if cursor_insert.fetchone():
                        errors.append(f"Row {row_num}: Email already exists")
                        continue
                    
                    hashed_password = generate_password_hash(password, method='scrypt')
                    
                    new_user_id = execute_insert_returning_id(db, cursor_insert, """
                        INSERT INTO users (first_name, last_name, email, password_hash, role)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (first_name, last_name, email, hashed_password, user_type))
                    
                    # Auto-assign free plan for clients
                    if user_type == 'client' and free_plan_id:
                        cursor_insert.execute("""
                            INSERT INTO subscriptions 
                            (user_id, plan_id, status, billing_cycle, price, start_date, end_date)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (new_user_id, free_plan_id, 'active', 'monthly', 0, 
                              datetime.now().date(), 
                              (datetime.now() + timedelta(days=30)).date()))
                    
                    created_count += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
        else:
            db.close()
            return jsonify({"error": "Invalid file format. Use CSV or XLSX"}), 400
        
        db.commit()
        db.close()
        
        return jsonify({
            "success": f"{created_count} users created",
            "created": created_count,
            "errors": errors
        }), 200
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/company/create", methods=["POST"])
def admin_create_company():
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        company_name = request.form.get("company_name")
        first_name = request.form.get("first_name")
        last_name = request.form.get("last_name")
        email = request.form.get("email")
        phone = request.form.get("phone")
        country = request.form.get("country")
        subscription_plan_id = request.form.get("subscription_plan_id")
        region = request.form.get("region")
        address = request.form.get("address")
        cnic = request.form.get("cnic")
        
        if not all([company_name, first_name, last_name, email, subscription_plan_id]):
            return jsonify({"error": "All fields required"}), 400
        
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if company email exists
        cursor.execute("SELECT id FROM companies WHERE email=%s", (email,))
        if cursor.fetchone():
            db.close()
            return jsonify({"error": "Company email already exists"}), 400

        cursor.execute("SELECT id FROM users WHERE email=%s", (email,))
        if cursor.fetchone():
            db.close()
            return jsonify({"error": "A user account with this email already exists"}), 400

        # Get subscription plan details
        cursor.execute("""
            SELECT id, monthly_price FROM subscription_plans 
            WHERE id=%s AND target_audience='company'
        """, (subscription_plan_id,))
        plan = cursor.fetchone()
        
        if not plan:
            db.close()
            return jsonify({"error": "Invalid subscription plan"}), 400
        
        # Use CNIC as password if provided, otherwise random
        password = cnic if cnic else os.urandom(12).hex()
        hashed_password = generate_password_hash(password, method='scrypt')

        # Handle profile image
        profile_img = None
        if 'profile_image' in request.files:
            profile_img = handle_file_upload(request.files['profile_image'], 'company', company_name)

        new_user_id = execute_insert_returning_id(db, cursor, """
            INSERT INTO users (username, email, password_hash, first_name, last_name, phone, country, role, is_active, cnic, profile_img)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'company', 1, %s, %s)
        """, (email.split("@", 1)[0], email, hashed_password, first_name, last_name, phone, country, cnic, profile_img))

        # Create company
        new_company_id = execute_insert_returning_id(db, cursor, """
            INSERT INTO companies (company_name, first_name, last_name, email, phone, country, profile_img)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (company_name, first_name, last_name, email, phone, country, profile_img))

        # Create subscription for company login account
        cursor.execute("""
            INSERT INTO subscriptions 
            (user_id, plan_id, status, billing_cycle, price, start_date, end_date)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (new_user_id, subscription_plan_id, 'active', 'monthly', plan['monthly_price'],
              datetime.now().date(), 
              (datetime.now() + timedelta(days=30)).date()))
        
        db.commit()
        db.close()
        
        return jsonify({"success": "Company created successfully"}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/company/update/<int:company_id>", methods=["POST"])
def admin_update_company(company_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        company_name = request.form.get("company_name")
        first_name = request.form.get("first_name")
        last_name = request.form.get("last_name")
        email = request.form.get("email")
        phone = request.form.get("phone")
        country = request.form.get("country")
        region = request.form.get("region")
        address = request.form.get("address")
        cnic = request.form.get("cnic")
        is_active = request.form.get("is_active") == "1" or request.form.get("is_active") == "true"
        
        if not all([company_name, first_name, last_name, email]):
            return jsonify({"error": "Required fields missing"}), 400
        
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if company exists
        cursor.execute("SELECT id FROM companies WHERE id=%s", (company_id,))
        if not cursor.fetchone():
            db.close()
            return jsonify({"error": "Company not found"}), 404

        # Check if email is already used by another company
        cursor.execute("SELECT id FROM companies WHERE email=%s AND id!=%s", (email, company_id))
        if cursor.fetchone():
            db.close()
            return jsonify({"error": "Email already used by another company"}), 400

        # Handle profile image if uploaded
        profile_img = None
        if 'profile_image' in request.files:
            profile_img = handle_file_upload(request.files['profile_image'], 'company', company_name)

        if profile_img:
            cursor.execute("""
                UPDATE companies 
                SET company_name=%s, first_name=%s, last_name=%s, email=%s, 
                    phone=%s, country=%s, region=%s, address=%s, profile_img=%s
                WHERE id=%s
            """, (company_name, first_name, last_name, email, phone, country, region, address, profile_img, company_id))

            cursor.execute("""
                UPDATE users 
                SET first_name=%s, last_name=%s, email=%s, phone=%s, country=%s, is_active=%s, cnic=%s, profile_img=%s
                WHERE email=(SELECT email FROM companies WHERE id=%s) AND role='company'
            """, (first_name, last_name, email, phone, country, is_active, cnic, profile_img, company_id))
        else:
            cursor.execute("""
                UPDATE companies 
                SET company_name=%s, first_name=%s, last_name=%s, email=%s, 
                    phone=%s, country=%s, region=%s, address=%s
                WHERE id=%s
            """, (company_name, first_name, last_name, email, phone, country, region, address, company_id))

            cursor.execute("""
                UPDATE users 
                SET first_name=%s, last_name=%s, email=%s, phone=%s, country=%s, is_active=%s, cnic=%s
                WHERE email=(SELECT email FROM companies WHERE id=%s) AND role='company'
            """, (first_name, last_name, email, phone, country, is_active, cnic, company_id))
        
        db.commit()
        db.close()
        
        return jsonify({"success": "Company updated successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/company/suspend/<int:company_id>", methods=["POST"])
def admin_suspend_company(company_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if company exists
        cursor.execute("SELECT id FROM companies WHERE id=%s", (company_id,))
        if not cursor.fetchone():
            db.close()
            return jsonify({"error": "Company not found"}), 404

        # Suspend company user account
        cursor.execute("""
            UPDATE users 
            SET is_active=0 
            WHERE email=(SELECT email FROM companies WHERE id=%s) AND role='company'
        """, (company_id,))
        
        db.commit()
        db.close()
        
        return jsonify({"success": "Company suspended successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/company/unsuspend/<int:company_id>", methods=["POST"])
def admin_unsuspend_company(company_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if company exists
        cursor.execute("SELECT id FROM companies WHERE id=%s", (company_id,))
        if not cursor.fetchone():
            db.close()
            return jsonify({"error": "Company not found"}), 404

        # Unsuspend company user account
        cursor.execute("""
            UPDATE users 
            SET is_active=1 
            WHERE email=(SELECT email FROM companies WHERE id=%s) AND role='company'
        """, (company_id,))
        
        db.commit()
        db.close()
        
        return jsonify({"success": "Company unsuspended successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/company/delete/<int:company_id>", methods=["DELETE"])
def admin_delete_company(company_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if company exists
        cursor.execute("SELECT id FROM companies WHERE id=%s", (company_id,))
        if not cursor.fetchone():
            db.close()
            return jsonify({"error": "Company not found"}), 404

        # Get company email for user deletion
        cursor.execute("SELECT email FROM companies WHERE id=%s", (company_id,))
        company = cursor.fetchone()
        
        # Delete company documents
        cursor.execute("DELETE FROM company_documents WHERE company_id=%s", (company_id,))
        
        # Delete company
        cursor.execute("DELETE FROM companies WHERE id=%s", (company_id,))
        
        # Delete associated user account
        if company:
            cursor.execute("DELETE FROM users WHERE email=%s AND role='company'", (company['email'],))
        
        db.commit()
        db.close()
        
        return jsonify({"success": "Company deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/subscription-plans/create", methods=["POST"])
def admin_create_subscription_plan():
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        name = request.form.get("name")
        target_audience = request.form.get("target_audience")
        monthly_price = float(request.form.get("monthly_price", 0))
        yearly_price = float(request.form.get("yearly_price", 0))
        features = request.form.get("features")
        
        if not all([name, target_audience, features]):
            return jsonify({"error": "Required fields missing"}), 400
        
        if target_audience not in ['individual', 'company']:
            return jsonify({"error": "Invalid target audience"}), 400
        
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("""
            INSERT INTO subscription_plans 
            (name, target_audience, monthly_price, yearly_price, features, is_active)
            VALUES (%s, %s, %s, %s, %s, 1)
        """, (name, target_audience, monthly_price, yearly_price, features))
        
        db.commit()
        db.close()
        
        log_activity(session["user_id"], f"Created subscription plan {name}", f"Target: {target_audience}")
        return jsonify({"success": "Subscription plan created successfully"}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/subscription-plans/update/<int:plan_id>", methods=["POST"])
def admin_update_subscription_plan(plan_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        name = request.form.get("name")
        target_audience = request.form.get("target_audience")
        monthly_price = float(request.form.get("monthly_price", 0))
        yearly_price = float(request.form.get("yearly_price", 0))
        features = request.form.get("features")
        
        if not all([name, target_audience, features]):
            return jsonify({"error": "Required fields missing"}), 400
        
        if target_audience not in ['individual', 'company']:
            return jsonify({"error": "Invalid target audience"}), 400
        
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if plan exists
        cursor.execute("SELECT id, name FROM subscription_plans WHERE id=%s", (plan_id,))
        plan = cursor.fetchone()
        if not plan:
            db.close()
            return jsonify({"error": "Subscription plan not found"}), 404

        # Update plan
        cursor.execute("""
            UPDATE subscription_plans 
            SET name=%s, target_audience=%s, monthly_price=%s, yearly_price=%s, features=%s
            WHERE id=%s
        """, (name, target_audience, monthly_price, yearly_price, features, plan_id))
        
        db.commit()
        db.close()
        
        log_activity(session["user_id"], f"Updated subscription plan {name}", f"Plan ID: {plan_id}")
        return jsonify({"success": "Subscription plan updated successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/subscription-plans/delete/<int:plan_id>", methods=["DELETE"])
def admin_delete_subscription_plan(plan_id):
    if "user_id" not in session or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Check if plan exists
        cursor.execute("SELECT id, name FROM subscription_plans WHERE id=%s", (plan_id,))
        plan = cursor.fetchone()
        if not plan:
            db.close()
            return jsonify({"error": "Subscription plan not found"}), 404

        # Check if plan is being used
        cursor.execute("SELECT COUNT(*) as count FROM subscriptions WHERE plan_id=%s", (plan_id,))
        usage = cursor.fetchone()
        if usage['count'] > 0:
            db.close()
            return jsonify({"error": "Cannot delete plan that is currently in use"}), 400

        # Delete plan
        cursor.execute("DELETE FROM subscription_plans WHERE id=%s", (plan_id,))
        
        db.commit()
        db.close()
        
        log_activity(session["user_id"], f"Deleted subscription plan {plan['name']}", f"Plan ID: {plan_id}")
        return jsonify({"success": "Subscription plan deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/subscription-plans/<target_audience>", methods=["GET"])
def get_subscription_plans(target_audience):
    try:
        if target_audience not in ['individual', 'company']:
            return jsonify({"error": "Invalid target audience"}), 400
        
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        cursor.execute("""
            SELECT id, name, monthly_price, yearly_price, features
            FROM subscription_plans
            WHERE target_audience=%s AND is_active=TRUE
            ORDER BY monthly_price ASC
        """, (target_audience,))
        
        plans = cursor.fetchall()
        db.close()
        
        return jsonify({"plans": plans}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/companies/users/import", methods=["POST"])
def company_users_import():
    if "user_id" not in session or session.get("role") != "company":
        return jsonify({"error": "Unauthorized"}), 403
        
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
            
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        cursor.execute("SELECT id, company_name FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
        comp = cursor.fetchone()
        
        if not comp:
            db.close()
            return jsonify({"error": "Company not found"}), 404
            
        company_id = comp["id"]
        company_name = "".join([c if c.isalnum() else "_" for c in comp["company_name"].lower()]).strip("_")
        
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'csv'
        filename = f"{company_name}_{company_id}.{ext}"
        
        upload_dir = os.path.join(UPLOAD_FOLDER, 'company_user')
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            
        filepath = os.path.join(upload_dir, filename)
        file.save(filepath)
        
        import pandas as pd
        if ext == 'csv':
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)
            
        df.columns = df.columns.str.strip()
        required_cols = ['FirstName', 'LastName', 'Username', 'Email', 'Phone', 'CNIC', 'Designation']
        for col in required_cols:
            if col not in df.columns:
                db.close()
                return jsonify({"error": f"Missing column: {col}"}), 400
                
        inserted_count = 0
        for _, row in df.iterrows():
            if pd.isna(row.get('Email')):
                continue
            
            first_name = str(row['FirstName']).strip() if not pd.isna(row['FirstName']) else ''
            last_name = str(row['LastName']).strip() if not pd.isna(row['LastName']) else ''
            username = str(row['Username']).strip() if not pd.isna(row['Username']) else ''
            email = str(row['Email']).strip()
            phone = str(row['Phone']).strip() if not pd.isna(row['Phone']) else ''
            cnic = str(row['CNIC']).strip() if not pd.isna(row['CNIC']) else ''
            designation = str(row['Designation']).strip() if not pd.isna(row['Designation']) else ''
            
            if not email:
                continue
                
            cursor.execute("SELECT id FROM users WHERE email=%s OR username=%s", (email, username))
            if cursor.fetchone():
                continue
                
            hashed_password = generate_password_hash(cnic if cnic else username, method='scrypt')
            
            new_user_id = execute_insert_returning_id(db, cursor, """
                INSERT INTO users (username, email, password_hash, first_name, last_name, company_id, role, cnic, designation, is_active)
                VALUES (%s, %s, %s, %s, %s, %s, 'company_user', %s, %s, 1)
            """, (username, email, hashed_password, first_name, last_name, company_id, cnic, designation))
            
            prefix = "".join(filter(str.isalnum, str(comp["company_name"])))[:2].upper()
            if len(prefix) < 2: prefix = (prefix + "XX")[:2]
            
            import string, random
            while True:
                random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
                employee_id = f"{prefix}{random_suffix}"
                cursor.execute("SELECT id FROM company_users WHERE employee_id=%s", (employee_id,))
                if not cursor.fetchone():
                    break
                    
            cursor.execute("""
                INSERT INTO company_users (user_id, company_id, employee_id)
                VALUES (%s, %s, %s)
            """, (new_user_id, company_id, employee_id))
            
            inserted_count += 1
            
        db.commit()
        db.close()
        
        return jsonify({"success": f"Imported {inserted_count} users successfully."})
        
    except Exception as e:
        print(f"Import Error: {e}")
        return jsonify({"error": f"Failed to process file: {str(e)}"}), 500

@app.route("/companies/users/delete/<int:user_id>", methods=["DELETE"])
def company_users_delete(user_id):
    if "user_id" not in session or session.get("role") != "company":
        return jsonify({"error": "Unauthorized"}), 403
        
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Verify the user to delete belongs to this company
        cursor.execute("SELECT id AS company_id FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
        comp = cursor.fetchone()
        
        if not comp:
            db.close()
            return jsonify({"error": "Not authorized"}), 403
            
        company_id = comp["company_id"]
        
        cursor.execute("SELECT id FROM users WHERE id=%s AND company_id=%s AND role='company_user'", (user_id, company_id))
        target_user = cursor.fetchone()
        
        if not target_user:
            db.close()
            return jsonify({"error": "User not found or unauthorized to delete"}), 404
            
        cursor.execute("DELETE FROM users WHERE id=%s", (user_id,))
        db.commit()
        db.close()
        
        return jsonify({"success": True})
        
    except Exception as e:
        print(f"Delete User Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/companies/users/create", methods=["POST"])
def company_create_user():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        if session.get("role") == "company":
            cursor.execute("SELECT id AS company_id FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
        else:
            cursor.execute("SELECT company_id FROM users WHERE id=%s", (session["user_id"],))
        user = cursor.fetchone()

        if not user or not user.get("company_id"):
            db.close()
            return jsonify({"error": "Not authorized"}), 403
        
        company_id = user["company_id"]
        
        first_name = request.form.get("first_name") or request.form.get("firstName")
        last_name = request.form.get("last_name") or request.form.get("lastName")
        email = request.form.get("email")
        password = request.form.get("password")
        cnic = request.form.get("cnic")
        designation = request.form.get("designation")
        username = request.form.get("username") or (email.split("@", 1)[0] if email else None)
        
        if not all([first_name, last_name, email, designation]):
            db.close()
            return jsonify({"error": "All fields required"}), 400

        password = password or cnic or os.urandom(12).hex()
        
        # Handle profile image
        profile_img = None
        if 'profile_image' in request.files:
            profile_img = handle_file_upload(request.files['profile_image'], 'company_user', f"{first_name}_{last_name}")
        
        # Check if user exists
        cursor.execute("SELECT id FROM users WHERE email=%s", (email,))
        if cursor.fetchone():
            db.close()
            return jsonify({"error": "Email already exists"}), 400
        
        hashed_password = generate_password_hash(password, method='scrypt')
        
        new_user_id = execute_insert_returning_id(db, cursor, """
            INSERT INTO users (username, first_name, last_name, email, password_hash, role, company_id, cnic, designation, profile_img)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (username, first_name, last_name, email, hashed_password, 'company_user', company_id, cnic, designation, profile_img))
        
        # Generate 8-character employee ID (First 2 chars of company name + 6 random alphanumeric)
        cursor.execute("SELECT company_name FROM companies WHERE id=%s", (company_id,))
        comp_record = cursor.fetchone()
        comp_name = comp_record["company_name"] if comp_record else "CO"
        prefix = "".join(filter(str.isalnum, comp_name))[:2].upper()
        if len(prefix) < 2: prefix = (prefix + "XX")[:2]
        
        import string, random
        while True:
            random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
            employee_id = f"{prefix}{random_suffix}"
            cursor.execute("SELECT id FROM company_users WHERE employee_id=%s", (employee_id,))
            if not cursor.fetchone():
                break
                
        cursor.execute("""
            INSERT INTO company_users (user_id, company_id, employee_id)
            VALUES (%s, %s, %s)
        """, (new_user_id, company_id, employee_id))
        
        db.commit()
        db.close()
        
        return jsonify({"success": "User created successfully"}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/companies/users/update/<int:user_id>", methods=["POST"])
def company_update_user(user_id):
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        # Permission check
        if session.get("role") == "company":
            cursor.execute("SELECT id AS company_id FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
        else:
            cursor.execute("SELECT company_id FROM users WHERE id=%s", (session["user_id"],))
        comp = cursor.fetchone()
        
        if not comp:
            db.close()
            return jsonify({"error": "Not authorized"}), 403
            
        company_id = comp["company_id"]
        
        # Verify user belongs to company
        cursor.execute("SELECT * FROM users WHERE id=%s AND company_id=%s", (user_id, company_id))
        target_user = cursor.fetchone()
        if not target_user:
            db.close()
            return jsonify({"error": "User not found or unauthorized"}), 404
            
        first_name = request.form.get("first_name") or request.form.get("firstName")
        last_name = request.form.get("last_name") or request.form.get("lastName")
        email = request.form.get("email")
        cnic = request.form.get("cnic")
        designation = request.form.get("designation")
        username = request.form.get("username")
        
        if not all([first_name, last_name, email, designation]):
            db.close()
            return jsonify({"error": "All fields required"}), 400
        
        # Handle profile image
        profile_img = None
        if 'profile_image' in request.files:
            profile_img = handle_file_upload(request.files['profile_image'], 'company_user', f"{first_name}_{last_name}")
            
        if profile_img:
            cursor.execute("""
                UPDATE users 
                SET first_name=%s, last_name=%s, email=%s, cnic=%s, designation=%s, username=%s, profile_img=%s
                WHERE id=%s
            """, (first_name, last_name, email, cnic, designation, username, profile_img, user_id))
        else:
            cursor.execute("""
                UPDATE users 
                SET first_name=%s, last_name=%s, email=%s, cnic=%s, designation=%s, username=%s
                WHERE id=%s
            """, (first_name, last_name, email, cnic, designation, username, user_id))
        
        db.commit()
        db.close()
        
        return jsonify({"success": "User updated successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/companies/users/import", methods=["POST"])
def company_import_users():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        if session.get("role") == "company":
            cursor.execute("SELECT id AS company_id FROM companies WHERE email=%s LIMIT 1", (session.get("email"),))
        else:
            cursor.execute("SELECT company_id FROM users WHERE id=%s", (session["user_id"],))
        user = cursor.fetchone()
        
        if not user or not user.get("company_id"):
            db.close()
            return jsonify({"error": "Not authorized"}), 403
        
        company_id = user["company_id"]
        
        file = request.files.get("file")
        if not file:
            db.close()
            return jsonify({"error": "No file provided"}), 400
        
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        
        created_count = 0
        errors = []
        
        if file_ext == 'csv':
            stream = io.TextIOWrapper(file.stream, encoding='utf-8')
            reader = csv.DictReader(stream)
            for row_num, row in enumerate(reader, start=2):
                try:
                    first_name = row.get('first_name', '').strip()
                    last_name = row.get('last_name', '').strip()
                    email = row.get('email', '').strip()
                    designation = row.get('designation', '').strip()
                    cnic = row.get('cnic', '').strip()
                    password = row.get('password', '').strip() or os.urandom(12).hex()
                    username = row.get('username', '').strip() or (email.split('@', 1)[0] if email else '')
                    
                    if not all([first_name, last_name, email, designation]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        continue
                    
                    cursor.execute("SELECT id FROM users WHERE email=%s", (email,))
                    if cursor.fetchone():
                        errors.append(f"Row {row_num}: Email already exists")
                        continue
                    
                    hashed_password = generate_password_hash(password, method='scrypt')
                    
                    cursor.execute("""
                        INSERT INTO users (username, first_name, last_name, email, password_hash, role, company_id, designation, cnic)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (username, first_name, last_name, email, hashed_password, 'company_user', company_id, designation, cnic))
                    
                    created_count += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
        
        elif file_ext == 'xlsx' and load_workbook:
            workbook = load_workbook(file.stream)
            worksheet = workbook.active
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    first_name = str(row[0] or '').strip()
                    last_name = str(row[1] or '').strip()
                    email = str(row[2] or '').strip()
                    designation = str(row[3] or '').strip()
                    cnic = str(row[4] or '').strip()
                    password = str(row[5] or '').strip() or os.urandom(12).hex()
                    username = email.split('@', 1)[0] if email else ''
                    
                    if not all([first_name, last_name, email, designation]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        continue
                    
                    cursor.execute("SELECT id FROM users WHERE email=%s", (email,))
                    if cursor.fetchone():
                        errors.append(f"Row {row_num}: Email already exists")
                        continue
                    
                    hashed_password = generate_password_hash(password, method='scrypt')
                    
                    cursor.execute("""
                        INSERT INTO users (username, first_name, last_name, email, password_hash, role, company_id, designation, cnic)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (username, first_name, last_name, email, hashed_password, 'company_user', company_id, designation, cnic))
                    
                    created_count += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
        else:
            db.close()
            return jsonify({"error": "Invalid file format. Use CSV or XLSX"}), 400
        
        db.commit()
        db.close()
        
        return jsonify({
            "success": f"{created_count} users created",
            "created": created_count,
            "errors": errors
        }), 200
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/settings/update", methods=["POST"])
def update_settings():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        user_id = session["user_id"]
        setting_type = request.form.get("type")
        
        db = get_db()
        cursor = db.cursor()
        
        if setting_type == "2fa":
            two_step_email = 1 if request.form.get("two_step_email") == "on" else 0
            two_step_sms = 1 if request.form.get("two_step_sms") == "on" else 0
            two_step_app = 1 if request.form.get("two_step_app") == "on" else 0
            
            cursor.execute("""
                UPDATE user_settings 
                SET two_step_email=%s, two_step_sms=%s, two_step_app=%s 
                WHERE user_id=%s
            """, (two_step_email, two_step_sms, two_step_app, user_id))
            
        elif setting_type == "notifications":
            promo_email = 1 if request.form.get("promo_email") == "on" else 0
            promo_sms = 1 if request.form.get("promo_sms") == "on" else 0
            
            cursor.execute("""
                UPDATE user_settings 
                SET promo_email=%s, promo_sms=%s 
                WHERE user_id=%s
            """, (promo_email, promo_sms, user_id))
            
        elif setting_type == "password":
            current_password = request.form.get("current_password")
            new_password = request.form.get("new_password")
            confirm_password = request.form.get("confirm_password")
            
            if not all([current_password, new_password, confirm_password]):
                return jsonify({"error": "All password fields are required"}), 400
                
            if new_password != confirm_password:
                return jsonify({"error": "New passwords do not match"}), 400
            
            cursor.execute("SELECT password_hash FROM users WHERE id=%s", (user_id,))
            user = cursor.fetchone()
            
            if not user or not check_password_hash(user[0], current_password):
                return jsonify({"error": "Incorrect current password"}), 400
            
            hashed_password = generate_password_hash(new_password, method='scrypt')
            cursor.execute("UPDATE users SET password_hash=%s WHERE id=%s", (hashed_password, user_id))
            cursor.execute("""
                INSERT INTO user_settings (user_id, password_last_changed) 
                VALUES (%s, CURRENT_TIMESTAMP) 
                ON CONFLICT (user_id) DO UPDATE SET password_last_changed = CURRENT_TIMESTAMP
            """, (user_id,))
            
        elif setting_type == "platform":
            if session.get("role") != "admin":
                return jsonify({"error": "Unauthorized"}), 403
            
            for key in request.form:
                if key.startswith("setting_"):
                    setting_key = key.replace("setting_", "")
                    setting_value = request.form.get(key)
                    cursor.execute("UPDATE settings SET setting_value=%s WHERE setting_key=%s", (setting_value, setting_key))
            
        db.commit()
        db.close()
        
        log_activity(user_id, f"Updated {setting_type} settings", activity_type=session.get("role", "client"))
        return jsonify({"success": f"{setting_type.capitalize()} settings updated successfully"}), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/process", methods=["POST"])
def process_interview():
    if "user_id" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    data = request.json
    provider = get_global_setting("ai_provider", "openai")
    model = get_global_setting("ai_model", "gpt-4")
    
    service = InterviewService(model_provider=provider, model_name=model)
    result, status = service.process_response(data)
    return jsonify(result), status

@app.route("/api/evaluate", methods=["POST"])
def evaluate_interview():
    if "user_id" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    data = request.json
    provider = get_global_setting("ai_provider", "openai")
    model = get_global_setting("ai_model", "gpt-4")
    
    service = InterviewService(model_provider=provider, model_name=model)
    result, status = service.evaluate_interview(data)
    
    # Save interview to database if successful
    if result.get("success"):
        try:
            db = get_db()
            cursor = get_db_cursor(db, dictionary=True)
            user_id = session["user_id"]
            role = data.get("role", "General")
            score = result.get("score", 0)
            feedback = result.get("overall", "")
            
            # For company users, get their company_id
            company_id = None
            cursor.execute("SELECT company_id FROM users WHERE id=%s", (user_id,))
            user_data = cursor.fetchone()
            if user_data:
                company_id = user_data.get("company_id")

            interview_type = 'company' if session.get('role') == 'company_user' else 'self'
            interview_id = data.get("interviewId")
            
            if interview_id:
                cursor.execute("""
                    UPDATE interviews 
                    SET overall_score = %s, status = 'completed', completed_at = CURRENT_TIMESTAMP 
                    WHERE id = %s AND user_id = %s
                """, (score, interview_id, user_id))
            else:
                cursor.execute("""
                    INSERT INTO interviews (user_id, target_role, overall_score, status, company_id, interview_type, completed_at) 
                    VALUES (%s, %s, %s, 'completed', %s, %s, CURRENT_TIMESTAMP)
                """, (user_id, role, score, company_id, interview_type))
                
            db.commit()
            db.close()
        except Exception as e:
            print(f"Error saving interview: {e}")
            
    return jsonify(result), status

@app.route("/api/notification/send", methods=["POST"])
def send_notification_api():
    if "user_id" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    data = request.json
    recipients = data.get("recipients", []) # List of user IDs or 'company_ID'
    title = data.get("title")
    message = data.get("message")
    sender_role = session.get("role")
    
    if not all([recipients, title, message]):
        return jsonify({"success": False, "error": "Missing required fields"}), 400
        
    try:
        db = get_db()
        cursor = get_db_cursor(db) # Standard cursor for writes
        
        target_user_ids = set()
        
        for r in recipients:
            r_str = str(r)
            if r_str.startswith("company_"):
                # Broadcast to all users of a company
                comp_id = r_str.replace("company_", "")
                cursor.execute("SELECT id FROM users WHERE company_id = %s", (comp_id,))
                for u in cursor.fetchall():
                    target_user_ids.add(u[0])
            else:
                # Specific user ID
                try:
                    target_user_ids.add(int(r))
                except ValueError:
                    continue # Skip invalid IDs
        
        # Insert notifications
        for uid in target_user_ids:
            cursor.execute("""
                INSERT INTO notifications (user_id, title, message, type)
                VALUES (%s, %s, %s, %s)
            """, (uid, title, message, sender_role if sender_role in ['admin', 'company'] else 'system'))
            
        db.commit()
        db.close()
        return jsonify({"success": True, "message": f"Notification sent to {len(target_user_ids)} recipients"}), 200
        
    except Exception as e:
        if 'db' in locals(): db.close()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/support/ticket/create", methods=["POST"])
def create_support_ticket():
    if "user_id" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    data = request.form
    user_id = session["user_id"]
    subject = data.get("subject")
    message = data.get("message")
    
    if not subject or not message:
        return jsonify({"success": False, "error": "Missing fields"}), 400
        
    try:
        db = get_db()
        cursor = db.cursor()
        ticket_number = f"TKT-{random.randint(10000, 99999)}"
        cursor.execute("""
            INSERT INTO support_tickets (user_id, ticket_number, subject, message, status)
            VALUES (%s, %s, %s, %s, 'open')
        """, (user_id, ticket_number, subject, message))
        db.commit()
        db.close()
        return jsonify({"success": True, "message": "Ticket created successfully"}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/support/ticket/reply", methods=["POST"])
def reply_support_ticket():
    if "user_id" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    data = request.json
    ticket_id = data.get("ticket_id")
    reply_text = data.get("reply")
    new_status = data.get("status")
    
    if not ticket_id or not reply_text:
        return jsonify({"success": False, "error": "Missing fields"}), 400
        
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Add reply
        cursor.execute("""
            INSERT INTO support_replies (ticket_id, user_id, reply_text, is_admin_reply)
            VALUES (%s, %s, %s, %s)
        """, (ticket_id, session["user_id"], reply_text, session.get("role") == "admin"))
        
        # Update ticket status if provided
        if new_status:
            cursor.execute("UPDATE support_tickets SET status = %s WHERE id = %s", (new_status, ticket_id))
            
        db.commit()
        db.close()
        return jsonify({"success": True, "message": "Reply sent successfully"}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/profile/update", methods=["POST"])
def update_profile():
    if "user_id" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    data = request.json
    user_id = session["user_id"]
    role = session.get("role")
    
    try:
        db = get_db()
        cursor = db.cursor()
        
        if role == 'company':
            # Update companies table
            cursor.execute("""
                UPDATE companies SET 
                company_name = %s, first_name = %s, last_name = %s, 
                phone = %s, country = %s, region = %s, address = %s
                WHERE email = %s
            """, (data.get('company_name'), data.get('first_name'), data.get('last_name'),
                  data.get('phone'), data.get('country'), data.get('region'), data.get('address'),
                  session.get('email')))
        else:
            # Update users table
            cursor.execute("""
                UPDATE users SET 
                first_name = %s, last_name = %s, phone = %s, 
                country = %s, region = %s, address = %s, 
                designation = %s, cnic = %s
                WHERE id = %s
            """, (data.get('first_name'), data.get('last_name'), data.get('phone'),
                  data.get('country'), data.get('region'), data.get('address'),
                  data.get('designation'), data.get('cnic'), user_id))
        
        db.commit()
        db.close()
        return jsonify({"success": True, "message": "Profile updated successfully"}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/profile/upload", methods=["POST"])
def upload_profile_img():
    if "user_id" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    if 'profile_image' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400
        
    file = request.files['profile_image']
    if file.filename == '':
        return jsonify({"success": False, "error": "No file selected"}), 400
        
    try:
        db = get_db()
        cursor = get_db_cursor(db, dictionary=True)
        
        cursor.execute("SELECT first_name, last_name, role FROM users WHERE id = %s", (session['user_id'],))
        user = cursor.fetchone()
        
        name_for_file = f"{user['first_name']}_{user['last_name']}"
        if user['role'] == 'company':
            cursor.execute("SELECT company_name FROM companies WHERE email = %s", (session.get('email'),))
            comp = cursor.fetchone()
            if comp: name_for_file = comp['company_name']
            
        img_url = handle_file_upload(file, user['role'], name_for_file)
        
        if session.get('role') == 'company':
            cursor.execute("UPDATE companies SET profile_img = %s WHERE email = %s", (img_url, session.get('email')))
            cursor.execute("UPDATE users SET profile_img = %s WHERE email = %s", (img_url, session.get('email')))
        else:
            cursor.execute("UPDATE users SET profile_img = %s WHERE id = %s", (img_url, session['user_id']))
        
        session["profile_image"] = img_url
            
        db.commit()
        db.close()
        return jsonify({"success": True, "profile_img": img_url}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# =========================
# LOGOUT
# =========================
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# =========================
# RESET TO DEFAULT PASSWORD
# =========================
@app.route("/api/users/reset-to-default/<int:user_id>", methods=["POST"])
def reset_to_default_password(user_id):
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 403
    
    # Only admin or company owner can reset passwords
    role = session.get("role")
    
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    
    # Check permission
    cursor.execute("SELECT * FROM users WHERE id=%s", (user_id,))
    target_user = cursor.fetchone()
    
    if not target_user:
        db.close()
        return jsonify({"error": "User not found"}), 404
        
    can_reset = False
    if role == "admin":
        can_reset = True
    elif role == "company":
        # Check if user belongs to this company
        cursor.execute("SELECT id FROM companies WHERE email=%s", (session.get("email"),))
        comp = cursor.fetchone()
        if comp and target_user["company_id"] == comp["id"]:
            can_reset = True
            
    if not can_reset:
        db.close()
        return jsonify({"error": "Insufficient permissions"}), 403
        
    if not target_user["cnic"]:
        db.close()
        return jsonify({"error": "User has no CNIC/ID Card number set as default"}), 400
        
    # Reset to CNIC
    hashed_password = generate_password_hash(target_user["cnic"], method='scrypt')
    cursor.execute("UPDATE users SET password_hash=%s WHERE id=%s", (hashed_password, user_id))
    
    db.commit()
    db.close()
    
    return jsonify({"success": "Password reset to default (ID Card number) successfully"}), 200

@app.route("/api/users/reset-by-email", methods=["POST"])
def reset_by_email_password():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 403
    
    data = request.get_json()
    email = data.get("email")
    cnic = data.get("cnic")
    
    if not email or not cnic:
        return jsonify({"error": "Email and CNIC are required"}), 400
        
    db = get_db()
    cursor = get_db_cursor(db, dictionary=True)
    
    # Check if user exists and role is allowed
    cursor.execute("SELECT * FROM users WHERE email=%s", (email,))
    user = cursor.fetchone()
    
    if not user:
        db.close()
        return jsonify({"error": "User not found"}), 404
        
    # Permission check
    can_reset = False
    if session.get("role") == "admin":
        can_reset = True
    elif session.get("role") == "company":
        # Check if user belongs to this company
        cursor.execute("SELECT id FROM companies WHERE email=%s", (session.get("email"),))
        comp = cursor.fetchone()
        if comp and user["company_id"] == comp["id"]:
            can_reset = True
            
    if not can_reset:
        db.close()
        return jsonify({"error": "Insufficient permissions"}), 403
        
    # Reset to CNIC
    hashed_password = generate_password_hash(cnic, method='scrypt')
    cursor.execute("UPDATE users SET password_hash=%s, cnic=%s WHERE id=%s", (hashed_password, cnic, user["id"]))
    
    db.commit()
    db.close()
    
    return jsonify({"success": "Password reset to default (ID Card number) successfully"}), 200

# =========================
# RUN
# =========================
if __name__ == "__main__":
    app.run(debug=True)
